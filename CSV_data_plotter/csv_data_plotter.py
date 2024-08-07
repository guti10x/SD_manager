######################################################################################################################################
#                                                                                                                                    #
#                           CSV DATA PLOTTER ANALYZER V1                                                                             # 
#                                    05/07/2024                                                                                      #
#                            By Daniel Gutiérrez Torres                                                                              #
#                                                                                                                                    #    
######################################################################################################################################

import csv
import sys
import os
from PyQt6.QtWidgets import QApplication, QHBoxLayout, QLabel, QLineEdit, QPushButton, QFileDialog, QCheckBox, QVBoxLayout, QWidget, QSizePolicy, QTableWidget, QTableWidgetItem, QSpacerItem, QScrollArea, QMessageBox, QFrame
from PyQt6.QtGui import QPixmap, QFont
from PyQt6.QtCore import Qt
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas, NavigationToolbar2QT
from matplotlib import ticker
import numpy as np
from statistics import mean, stdev
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
from matplotlib import pyplot as plt  

# TOOLBAR GRÁFICAS
class CustomToolbar(NavigationToolbar2QT):
    def __init__(self, canvas, parent=None):
        super().__init__(canvas, parent)
        # Estilo de la barra de herramientas
        self.setStyleSheet(
            "background-color: #FFFFFF;"  # Color de fondo blanco
            "QToolButton {"
            "    color: white;"  # Color del texto en los botones
            "    background-color: #111111;"  # Fondo de los botones
            "}"
            "QToolButton:hover {"
            "    background-color: #FFFFFF;"  # Fondo de los botones al pasar el mouse
            "}"
            "QToolButton:pressed {"
            "    background-color: #FFFFFF;"  # Fondo de los botones al presionar
            "}"
        )
        self.hide_home_button()

    def hide_home_button(self):
        # Iterate through the toolbar's children to find the "Home" button
        for action in self.actions():
            if action.text() == 'Home':
                action.setVisible(False)
                break

# VENTANA MAIN
class Ventana(QWidget):

    # SETUP##### #############################################################################################################################
    def __init__(self):
        super().__init__()

        # equivalencias ID <-> parámetro
        self.ID_TO_PARAM = {
            1: 'Engine temp',
            2: 'Batery vol',
            3: 'Brake 1 temp',
            4: 'Brake 2 temp',
            5: 'Brake 3 temp',
            6: 'Brake 4 temp',
            7: 'Gear',
            8: 'Speed',
        }

        self.graficas = {} 

        self.initUI()

    # INTERFAZ ###############################################################################################################################
    def initUI(self):

        # Configuración de la ventana
        self.setGeometry(50, 50, 1100, 500) 
        self.setWindowTitle('FUEM CSV data plotter analyzer')

        # PLANTILLA DE LA VENTANA #################################################################################################
        layout = QVBoxLayout()
        self.setLayout(layout)

        # PLANTILLA HEADER INPUT + OPCIONES #################################################################################
        # Crear el layout header_menu
        # Create the header_menu_widget and layout_header_menu
        self.header_menu_widget = QWidget()
        self.layout_header_menu = QVBoxLayout(self.header_menu_widget)

        # Add the header menu widget to the main layout
        layout.addWidget(self.header_menu_widget)

        # Add the header menu widget to the main layout
        layout.addWidget(self.header_menu_widget)

        # TÍTULO ###################################################################################################
        titulo_label = QLabel('CSV Data Plotter Analyzer', self)
        titulo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)  
        titulo_label.setFont(QFont('Helvetica Neue', 24, QFont.Weight.Bold))
        titulo_label.setStyleSheet("border:None;")  
        self.layout_header_menu.addWidget(titulo_label, Qt.AlignmentFlag.AlignHCenter)

        # Spacer para añadir espacio debajo del título
        spacer1 = QSpacerItem(0, 10, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.layout_header_menu.addItem(spacer1)

        # INPUT RUTA CSV ##########################################################################################
        input_ruta = QHBoxLayout()

        input_ruta.addSpacing(60)

        # Crear y añadir el QLabel
        input_label = QLabel('Path to the CSV file:', self)
        input_label.setStyleSheet("font-size: 17px; border:None;")
        input_ruta.addWidget(input_label, alignment=Qt.AlignmentFlag.AlignLeft)

        # Crear y añadir el QLineEdit
        self.ruta_input = QLineEdit(self)
        self.ruta_input.setPlaceholderText("Enter file path")
        self.ruta_input.setStyleSheet("padding: 5px; font-size: 14px; border: 1px solid #ccc; border-radius: 4px;")
        self.ruta_input.setAlignment(Qt.AlignmentFlag.AlignLeft)
        input_ruta.addWidget(self.ruta_input)

        # Crear y añadir el QPushButton
        btn_subir_csv = QPushButton('Select CSV File', self)
        btn_subir_csv.clicked.connect(self.abrir_dialogo_csv)
        btn_subir_csv.setStyleSheet(
            "background-color: #000000; color: white; padding: 6px 20px; border: 2px solid #000000; border-radius: 5px; "
            "font-size: 14px; text-align: center;"
            "transition: background-color 0.3s, transform 0.2s;"
            "}"
            "QPushButton:hover {"
            "background-color: #BF3F29;"
            "border-color: #FFFFFF;" 
            "}"
            "QPushButton:pressed {"
            "background-color: #A7301C;"
            "transform: scale(0.95);"
            "border-color: #FFFFFF;"  
            "}"
        )
        input_ruta.addWidget(btn_subir_csv) 

        input_ruta.addSpacing(50)

        # Añadir el layout horizontal al layout_header_menu
        self.layout_header_menu.addLayout(input_ruta)

        # Spacer para añadir espacio debajo de los botones
        spacer2 = QSpacerItem(0, 15, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.layout_header_menu.addItem(spacer2)

        # BOTONES OPCIONES #########################################################################################
        # Layout horizontal para los botones
        layout_botones = QHBoxLayout()

        # Espacio entre los botones
        layout_botones.addSpacing(170)

        # Botón para cargar los datos del archivo
        btn_cargar_datos = QPushButton('Visualize Data', self)
        btn_cargar_datos.clicked.connect(self.cargar_datos)
        btn_cargar_datos.setStyleSheet(
            "background-color: #E41B12; color: white; padding: 10px 20px; border-radius: 5px; "
            "font-size: 14px; border: none; text-align: center;"
            "transition: background-color 0.3s, transform 0.2s;"
            "}"
            "QPushButton:hover {"
            "background-color: #BF3F29;"
            "}"
            "QPushButton:pressed {"
            "background-color: #A7301C;"
            "transform: scale(0.95);"
            "}")
        layout_botones.addWidget(btn_cargar_datos)

        # Espacio entre los botones
        layout_botones.addSpacing(50)

        # Botón para comparar datos 
        btn_compare = QPushButton('Compare Data', self)
        btn_compare.clicked.connect(self.compare_data)
        btn_compare.setStyleSheet(
            "background-color: #E41B12; color: white; padding: 10px 20px; border-radius: 5px; "
            "font-size: 14px; border: none; text-align: center;"
            "transition: background-color 0.3s, transform 0.2s;"
            "}"
            "QPushButton:hover {"
            "background-color: #BF3F29;"
            "}"
            "QPushButton:pressed {"
            "background-color: #A7301C;"
            "transform: scale(0.95);"
            "}") 
        layout_botones.addWidget(btn_compare)

        # Espacio entre los botones
        layout_botones.addSpacing(50)

        # Botón para generar informe en Excel
        btn_excel = QPushButton('Generate Excel', self)
        btn_excel.clicked.connect(self.show_excel_report)
        btn_excel.setStyleSheet(
            "background-color: #E41B12; color: white; padding: 10px 20px; border-radius: 5px; "
            "font-size: 14px; border: none; text-align: center;"
            "transition: background-color 0.3s, transform 0.2s;"
            "}"
            "QPushButton:hover {"
            "background-color: #BF3F29;"
            "}"
            "QPushButton:pressed {"
            "background-color: #A7301C;"
            "transform: scale(0.95);"
            "}")  # Botón rojo con efectos de hover y click
        layout_botones.addWidget(btn_excel)

        # Espacio entre los botones
        layout_botones.addSpacing(50)

        # Botón para generar informe en PDF
        btn_pdf = QPushButton('Generate PDF', self)
        btn_pdf.clicked.connect(self.generate_pdf_report)
        btn_pdf.setStyleSheet(
            "background-color: #E41B12; color: white; padding: 10px 20px; border-radius: 5px; "
            "font-size: 14px; border: none; text-align: center;"
            "transition: background-color 0.3s, transform 0.2s;"
            "}"
            "QPushButton:hover {"
            "background-color: #BF3F29;"
            "}"
            "QPushButton:pressed {"
            "background-color: #A7301C;"
            "transform: scale(0.95);"
            "}")  # Botón rojo con efectos de hover y click
        layout_botones.addWidget(btn_pdf)

        # Espacio entre los botones
        layout_botones.addSpacing(120)

        # Añadir el layout horizontal de los botones al layout_header_menu
        self.layout_header_menu.addLayout(layout_botones)
        self.layout_header_menu.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Spacer para añadir espacio debajo de los botones
        spacer3 = QSpacerItem(0, 18, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.layout_header_menu.addItem(spacer3)

        self.layout_header_menu.setContentsMargins(10, 10, 10, 10)  # Margenes del header

        # Crear y configurar el widget del header
        header_widget = QWidget()
        header_widget.setLayout(self.layout_header_menu)
        header_widget.setStyleSheet(
            "background-color: rgba(245, 245, 245, 200); border: 2px solid #E41B12; "
            "border-radius: 30px; padding: 10px; "
            "box-shadow: 0px 2px 5px rgba(0, 0, 0, 0.2);"
        )
        # Configurar el tamaño del header
        header_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        header_widget.setFixedHeight(250)  # Ajusta la altura según necesidad

        # Añadir el widget del header al layout principal
        layout.addWidget(header_widget, Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignTop)

        # Configurar el layout principal para centrar su contenido
        self.setLayout(layout)


        # LAYOUTS PARA AJUSTES DE FUNCIONALIDADES DE CADA BOTÓN ###################################################

        # Contenedor para los elementos del generador de Excel
        self.excel_container_frame = QFrame(self)  
        self.excel_container_frame.setFixedWidth(1450)
        self.excel_container_layout = QVBoxLayout(self.excel_container_frame)

        # Estilos para el contenedor de Excel
        self.excel_container_frame.setStyleSheet(""" 
            QFrame {
                border: 2px solid #FF0000; 
                border-radius: 8px;       
                padding: 10px;           
                background-color: #FFFFFF;
            }
        """)

        self.excel_container_layout.setContentsMargins(10, 10, 10, 10)
        self.excel_container_layout.setSpacing(10)

        # Sub-layout para el nombre del archivo ##########################################################
        head_layout = QHBoxLayout()
        head_layout.setContentsMargins(0, 0, 0, 0)  # Sin márgenes
        head_layout.setSpacing(10)  # Espacio entre los elementos

        # Etiqueta para el nombre del desplegable de configuración para crear Excel
        file_name_label = QLabel('Generate Excel Settings', self)
        file_name_label.setStyleSheet("font-size: 25px; margin:0; border: none; font-weight: bold; color: #333;")
        head_layout.addWidget(file_name_label)

        #Botón para cerrar desplegable de configuración para crear Excel
        self.close_button = QPushButton('X', self)
        self.close_button.clicked.connect(self.hide_excel_report)
        self.close_button.setStyleSheet("padding: 5px 10px; font-size: 14px; background-color: #E41B12; color: white; border: none; border-radius: 4px;")
        head_layout.addWidget(self.close_button, alignment=Qt.AlignmentFlag.AlignRight)

        self.excel_container_layout.addLayout(head_layout)

        # Sub-layout para el nombre del archivo ##########################################################
        file_input_layout = QHBoxLayout()
        file_input_layout.setContentsMargins(0, 0, 0, 0)  # Sin márgenes
        file_input_layout.setSpacing(10)  # Espacio entre los elementos

        # Etiqueta para el nombre del archivo
        file_name_label = QLabel('Filename:', self)
        file_name_label.setStyleSheet("font-size: 16px; margin:0; border: none; font-weight: bold; color: #333;")
        file_input_layout.addWidget(file_name_label)

        # Campo de texto para ingresar el nombre del archivo
        self.file_name_input = QLineEdit(self)
        self.file_name_input.setPlaceholderText("Enter a file name")
        self.file_name_input.setStyleSheet("padding: 5px; font-size: 14px; border: 1px solid #ccc; border-radius: 4px;")
        file_input_layout.addWidget(self.file_name_input)

        # Añadir el sub-layout de entrada al contenedor de Excel
        self.excel_container_layout.addLayout(file_input_layout)

        # Checkbox para añadir fecha al nombre ##########################################################
        checkbox_layout = QHBoxLayout()
        checkbox_layout.setSpacing(10)

        self.timestamp_checkbox = QCheckBox("Add current date to the filename", self)
        self.timestamp_checkbox.setStyleSheet("margin-left:15px; font-size: 16px; color: #555;")
        checkbox_layout.addWidget(self.timestamp_checkbox)

        # Añadir el sub-layout del checkbox al contenedor de Excel
        self.excel_container_layout.addLayout(checkbox_layout)

        # Sub-layout para el botón de generación del excell ##############################################
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        # Botón para generar el archivo Excel
        self.generate_button = QPushButton('Generate Excel', self)
        self.generate_button.clicked.connect(self.generate_excel_report)
        self.generate_button.setStyleSheet("padding: 10px 20px; font-size: 14px; background-color: #E41B12; color: white; border: none; border-radius: 4px; margin-left: 15px; margin-top:20px;")

        button_layout.addWidget(self.generate_button, alignment=Qt.AlignmentFlag.AlignLeft)

        # Añadir el sub-layout del botón al contenedor de Excel
        self.excel_container_layout.addLayout(button_layout)

        # Sub-layout mensajes de error/exito ###############################################################
        self.message_label_xlsx = QLabel('', self)
        self.message_label_xlsx.setStyleSheet("font-size: 16px; margin:0; border: none; color: #333;")
        self.message_label_xlsx.hide()
        self.excel_container_layout.addWidget(self.message_label_xlsx, alignment=Qt.AlignmentFlag.AlignLeft)

        # Añadir el contenedor de Excel al layout principal
        layout.addWidget(self.excel_container_frame, alignment=Qt.AlignmentFlag.AlignCenter)
        self.toggle_excel_container(False)


        # LAYOUT GRÁFICAS INDIVIDUALES Y TABLAS ####################################################################

        # Widget scrollable para las tablas
        self.scroll_content = QWidget()
        self.scroll_content.setFixedWidth(1465)
        self.layout_tablas = QVBoxLayout(self.scroll_content)
        
        # Crear un área de scroll
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setWidget(self.scroll_content)
        self.scroll_area.setFixedHeight(520)
        self.scroll_area.setFixedWidth(1510)
        self.scroll_area.hide()
        self.scroll_area.setStyleSheet(
            "QScrollArea {"
            "background-color: rgba(245, 245, 245, 200);"
            "border: 2px solid #E41B12;"
            "border-radius: 10px;"
            "padding: 10px;"
            "}"
            "QScrollArea > QWidget > QWidget {"
            "background-color: rgba(245, 245, 245, 200);"
            "border-radius: 10px;"
            "}"
        )

        # Agregar el área de scroll al layout principal
        layout.addWidget(self.scroll_area)

        # LAYOUT GRÁFICA CON SUBPLOTS #############################################################################
        self.subplot_area = QWidget()
        self.subplot_area.setFixedHeight(520)
        self.subplot_area.hide()
        layout.addWidget(self.subplot_area)

        # IMG fondo
        self.backgroundLogo = QLabel(self)
        self.backgroundLogo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ruta_script = os.path.dirname(os.path.abspath(__file__))  
        ruta_imagen = os.path.join(ruta_script, 'assets/IMG_fondo.PNG') 
        pixmap = QPixmap(ruta_imagen)  
        pixmap = pixmap.scaled(self.size(), Qt.AspectRatioMode.IgnoreAspectRatio)
        self.backgroundLogo.setPixmap(pixmap)
        layout.addWidget(self.backgroundLogo)

    ruta_to_input = None

    # FUNCIONES ##############################################################################################################################

    # Abrir un diálogo para seleccionar archivo CSV
    def abrir_dialogo_csv(self):
        archivo_csv, _ = QFileDialog.getOpenFileName(self, "Seleccionar archivo CSV", "", "Archivos CSV (*.csv)")
        
        if archivo_csv:
            self.ruta_to_input = archivo_csv
            self.ruta_input.setText(archivo_csv)
            #print(f"Ruta del archivo seleccionado: {self.ruta_input}")

    # Cerrar todos los desplegables desplegados 
    def close_desplegables(self):
        self.scroll_area.setVisible(False)
        self.subplot_area.setVisible(False)
        self.toggle_excel_container(False)
        
    # Obtener la fecha y hora actuales
    def get_current_timestamp(self):
        now = datetime.now()
        timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
        return timestamp

    # Mostrar u ocultar el contenedor de Excel
    def toggle_excel_container(self, visible):
        self.excel_container_frame.setVisible(visible)

    # Mostrar el contenedor de Excel
    def show_excel_report(self):
        self.toggle_excel_container(True)

    # Ocultar el contenedor de Excel
    def hide_excel_report(self):
        self.toggle_excel_container(False)

    #Generar plot de graficas comparativas
    def generate_excel_report(self):
        
        # Cerrar todos los desplegables abiertos
        self.close_desplegables()

        # Asegúrate de que el mensaje de error se oculte inicialmente
        self.message_label_xlsx.hide()

        if self.ruta_to_input is None:
            self.message_label_xlsx.setText("¡Selecciona un archivo CSV!")
            self.message_label_xlsx.setStyleSheet("color: #FF0000; border: none; margin:0; font-size: 16px;")
            self.message_label_xlsx.show()
            return 

        file_name = self.file_name_input.text().strip()  

        # Verificar si el campo de texto está vacío
        if not file_name:
            self.message_label_xlsx.setText("¡Nombre del excel a generar no inicializado !")
            self.message_label_xlsx.setStyleSheet("color: #FF0000; border: none; margin:0; font-size: 16px;")
            self.message_label_xlsx.show()
            return 
        
        # Verificar el estado del checkbox
        if self.timestamp_checkbox.isChecked():
            file_name = f'{file_name}_{self.get_current_timestamp()}'
        else:
            file_name = f'{file_name}'
          
        try:
            # Leer el archivo CSV
            self.csv_data = pd.read_csv(self.ruta_to_input, header=None, names=['ID', 'Value', 'Time'])
            
            with pd.ExcelWriter(f'{file_name}.xlsx', engine='openpyxl') as writer:
                # Bandera para verificar si se añadió al menos una hoja
                sheet_added = False
                
                for param_id, param_name in self.ID_TO_PARAM.items():
                    # Filtrar los datos por ID
                    df_filtered = self.csv_data[self.csv_data['ID'] == param_id]
                    if not df_filtered.empty:
                        # Renombrar la columna 'Value' a nombre del parámetro
                        df_filtered = df_filtered[['Value', 'Time']]
                        df_filtered.rename(columns={'Value': param_name}, inplace=True)
                        # Escribir en la hoja correspondiente
                        df_filtered.to_excel(writer, sheet_name=param_name, index=False)
                        sheet_added = True
                
                # Si no se ha añadido ninguna hoja, agregar una hoja predeterminada
                if not sheet_added:
                    pd.DataFrame({'Message': ['No data available for the given parameters']}).to_excel(writer, sheet_name='No Data', index=False)

            # Abrir el archivo de nuevo para modificar el estilo
            workbook = load_workbook('output.xlsx')
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            workbook.save('output.xlsx')

            # Mostrar mensaje de éxito
            self.message_label_xlsx.setText("¡Archivo generado exitosamente!")
            self.message_label_xlsx.setStyleSheet("color: #4CAF50; border: none; margin:0; font-size: 16px;")
            self.message_label_xlsx.show()

        except Exception as e:
            # Mostrar mensaje de error en caso de excepción
            self.message_label_xlsx.setText(f"Error: {str(e)}")
            self.message_label_xlsx.setStyleSheet("color: #FF0000; border: none; margin:0; font-size: 16px;")
            self.message_label_xlsx.show()
    
    def compare_data(self):
        # Cerrar todos los desplegables abiertos
        self.close_desplegables()

        file_name = self.ruta_to_input

        # Ensure that the subplot_area has a layout
        if not self.subplot_area.layout():
            self.subplot_area.setLayout(QVBoxLayout())

        # Clear the layout of subplot_area
        for i in reversed(range(self.subplot_area.layout().count())):
            widget = self.subplot_area.layout().itemAt(i).widget()
            if widget:
                widget.deleteLater()

        if not file_name:
            title_subplot_compare = QLabel("No data to display")
            title_subplot_compare.setAlignment(Qt.AlignmentFlag.AlignCenter)
            title_subplot_compare.setStyleSheet("font-size: 11pt; font-weight: bold; margin: 10px;")
            self.subplot_area.layout().addWidget(title_subplot_compare)
            self.subplot_area.show()
            return

        # Create a list of IDs to display
        ids_a_mostrar = list(self.ID_TO_PARAM.keys())

        # Create a dictionary to store data by ID
        datos_por_id = {id_: [] for id_ in ids_a_mostrar}

        with open(file_name, 'r', newline='') as csv_file:
            csv_reader = csv.reader(csv_file)
            next(csv_reader)
            for line in csv_reader:
                if len(line) >= 3:
                    id_ = int(line[0])
                    valor = line[1]
                    timestamp = line[2]
                    if id_ in datos_por_id:
                        datos_por_id[id_].append((timestamp, valor))

        # Define figure size (width, height)
        fig, axs = plt.subplots(len(ids_a_mostrar), 1, sharex=True, figsize=(12, 10))

        for i, id_ in enumerate(ids_a_mostrar):
            x_data = [timestamp.split()[1] for timestamp, _ in datos_por_id[id_]]  # Extract only the time part
            y_data = [float(valor) for _, valor in datos_por_id[id_]]

            axs[i].plot(x_data, y_data, marker='o', color='red')  # Set line color to red

            # Remove the default y-axis label
            axs[i].set_ylabel("")

            # Add custom styled text as y-axis label
            axs[i].text(-0.1, 0.5, self.ID_TO_PARAM[id_], 
            transform=axs[i].transAxes, 
            fontsize=10, 
            verticalalignment='center', 
            horizontalalignment='right',
            bbox=dict(facecolor='red', edgecolor='black', boxstyle='round,pad=0.5'))

            axs[i].tick_params(axis='x', which='both', bottom=False, labelbottom=False)  # Hide x-axis labels for subplots
            axs[i].grid(True)

        axs[-1].tick_params(axis='x', which='both', bottom=True, labelbottom=True)  # Show x-axis labels only on the last subplot
        axs[-1].xaxis.set_major_locator(ticker.MaxNLocator(nbins=15))  # Num marcas eje X
        plt.setp(axs[-1].get_xticklabels(), rotation=0, ha='right')  # Rotación etiquetas eje X
        axs[-1].set_xlabel('Time')

        # Adjust the layout to remove margins and fit content
        fig.tight_layout()

        # Create FigureCanvas and add it to the layout of subplot_area
        canvas = FigureCanvas(fig)
        toolbar = CustomToolbar(canvas, self.subplot_area)

        # Add the toolbar and canvas to the layout
        self.subplot_area.layout().addWidget(toolbar)
        self.subplot_area.layout().addWidget(canvas)

        # Show the subplot area
        self.subplot_area.show()

    def generate_pdf_report(self):
        
        # Cerrar todos los desplegables abiertos
        self.close_desplegables()

    def cargar_datos(self):

        self.header_menu_widget.hide()

        file_name = self.ruta_to_input

        # Limpiar el layout existente antes de agregar nuevos elementos
        for i in reversed(range(self.layout_tablas.count())):
            widget = self.layout_tablas.itemAt(i).widget()
            if widget:
                widget.deleteLater()

        # Crear una lista de IDs que queremos mostrar
        ids_a_mostrar = list(self.ID_TO_PARAM.keys())

        # Crear un diccionario para almacenar los datos por ID
        datos_por_id = {id_: [] for id_ in ids_a_mostrar}

        if file_name:
            with open(file_name, 'r', newline='') as csv_file:
                csv_reader = csv.reader(csv_file)
                next(csv_reader)  # Saltar la cabecera si la hay
                for line in csv_reader:
                    if len(line) >= 3:  # Asegurar que la línea tenga al menos 3 elementos (ID, Valor, Timestamp)
                        id_ = int(line[0])  # Suponiendo que el ID está en la primera columna y es un entero
                        valor = line[1]
                        timestamp = line[2]

                        # Formatear el timestamp: eliminar día, mes, año y dejar solo hora, minutos y segundos
                        try:
                            dt = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                            formatted_timestamp = dt.strftime('%H:%M:%S')
                        except ValueError:
                            # Si el formato es incorrecto, devolver el timestamp original
                            formatted_timestamp = timestamp

                        if id_ in datos_por_id:
                            datos_por_id[id_].append((formatted_timestamp, valor))

            # Recorrer los IDs y crear una tabla para cada uno
            for id_ in ids_a_mostrar:
                titulo = QLabel(self.ID_TO_PARAM[id_])
                titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
                titulo.setStyleSheet("font-size: 16pt; font-weight: bold; margin: 10px;")
                self.layout_tablas.addWidget(titulo)

                # Crear un widget para contener el layout horizontal
                h_widget = QWidget()

                # Crear un layout horizontal para la tabla y la gráfica
                h_layout = QHBoxLayout(h_widget)

                # Crear la tabla
                tabla = QTableWidget()
                tabla.setColumnCount(2)
                tabla.setHorizontalHeaderLabels(['Time (x)', 'Value (y)'])

                data = datos_por_id[id_]
                tabla.setRowCount(len(data))

                for row, (timestamp, valor) in enumerate(data):
                    item_timestamp = QTableWidgetItem(timestamp)
                    item_valor = QTableWidgetItem(valor)

                    # Centrar el texto de las celdas
                    item_timestamp.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item_valor.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                    tabla.setItem(row, 0, item_timestamp)
                    tabla.setItem(row, 1, item_valor)

                # Ajustar el ancho de la columna 'Timestamp'
                tabla.setColumnWidth(0, 200)

                # Establecer el alto y ancho de la tabla
                altura_tabla = 560
                ancho_tabla = 242
                tabla.setFixedHeight(altura_tabla)
                tabla.setFixedWidth(ancho_tabla)

                h_layout.addWidget(tabla)

                # Widget para contener la barra de herramientas y la gráfica
                graph_widget = QWidget()
                graph_layout = QVBoxLayout(graph_widget)

                # GRÁFICA
                # Crear la gráfica asociada
                x_data = [timestamp for timestamp, _ in data]
                y_data = [float(valor) for _, valor in data]

                # Ajustar el tamaño de la gráfica 
                fig = Figure(figsize=(8, 6))  
                ax = fig.add_subplot(111)
                ax.plot(x_data, y_data, marker='o', color='#E41B12') 

                ax.set_xlabel('Time')
                ax.set_ylabel('Value')

                # Ajustar las etiquetas del eje X para evitar solapamiento
                ax.xaxis.set_major_locator(ticker.MaxNLocator(nbins=15))  # Num etiquetas en el eje X
                for label in ax.get_xticklabels():
                    label.set_rotation(30)  # Rotación etiquetas en el eje X
                    label.set_horizontalalignment('right')

                # Crear el canvas para la gráfica
                canvas = FigureCanvas(fig)

                # Añadir la barra de herramientas personalizada de matplotlib
                toolbar = CustomToolbar(canvas, self)
                graph_layout.addWidget(toolbar)  # Añadir la barra de herramientas al layout vertical
                graph_layout.addWidget(canvas)   # Añadir la gráfica al layout vertical

                # Ajustar el ancho y alto del widget que contiene la gráfica
                graph_widget.setFixedWidth(1150)  # Ajustar el ancho según tus necesidades
                graph_widget.setFixedHeight(580)  # El mismo alto que la tabla

                # Agregar el widget con la gráfica y la barra de herramientas al layout horizontal
                h_layout.addWidget(graph_widget)

                # Ajustar el ancho del widget que contiene el layout horizontal
                h_widget.setFixedWidth(1500)  # Ajustar el ancho total según tus necesidades

                # Agregar el widget con el layout horizontal al layout principal vertical
                self.layout_tablas.addWidget(h_widget)

        else:
            # Crear etiqueta de título
            titulo = QLabel("No data to display")
            titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
            titulo.setStyleSheet("font-size: 18pt; font-weight: bold; margin: 10px;")
            self.layout_tablas.addWidget(titulo)

        # Mostrar el QScrollArea después de cargar los datos
        self.scroll_area.show()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ventana = Ventana()
    ventana.showMaximized() 
    ventana.show()
    sys.exit(app.exec())